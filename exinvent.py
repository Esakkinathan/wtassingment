import tkinter as tk
from tkinter import messagebox
import openpyxl

class SupermarketInventory:
    def __init__(self):
        self.inventory = {}
        self.load_inventory()

    def load_inventory(self):
        try:
            workbook = openpyxl.load_workbook("inventory.xlsx")
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                product_id, name, price, quantity = row
                self.inventory[product_id] = {
                    'name': name,
                    'price': price,
                    'quantity': quantity
                }
            workbook.close()
        except FileNotFoundError:
            self.inventory = {}

    def save_inventory(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Product ID", "Name", "Price", "Quantity"])
        for product_id, details in self.inventory.items():
            sheet.append([product_id, details['name'], details['price'], details['quantity']])
        workbook.save("inventory.xlsx")

    def add_product(self, product_id, name, price, quantity):
        if product_id in self.inventory:
            self.inventory[product_id]['quantity'] += quantity
        else:
            self.inventory[product_id] = {
                'name': name,
                'price': price,
                'quantity': quantity
            }
        self.save_inventory()

    def remove_product(self, product_id, quantity):
        if product_id in self.inventory:
            if self.inventory[product_id]['quantity'] >= quantity:
                self.inventory[product_id]['quantity'] -= quantity
                self.save_inventory()
            else:
                messagebox.showerror("Error", "Not enough stock available for this product.")
        else:
            messagebox.showerror("Error", "Product not found in inventory.")

    def view_inventory(self):
        return self.inventory

class InventoryManagementApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Supermarket Inventory Management System")

        self.supermarket = SupermarketInventory()

        self.product_id_label = tk.Label(root, text="Product ID:")
        self.product_id_label.grid(row=0, column=0)
        self.product_id_entry = tk.Entry(root)
        self.product_id_entry.grid(row=0, column=1)

        self.name_label = tk.Label(root, text="Name:")
        self.name_label.grid(row=1, column=0)
        self.name_entry = tk.Entry(root)
        self.name_entry.grid(row=1, column=1)

        self.price_label = tk.Label(root, text="Price:")
        self.price_label.grid(row=2, column=0)
        self.price_entry = tk.Entry(root)
        self.price_entry.grid(row=2, column=1)

        self.quantity_label = tk.Label(root, text="Quantity:")
        self.quantity_label.grid(row=3, column=0)
        self.quantity_entry = tk.Entry(root)
        self.quantity_entry.grid(row=3, column=1)

        self.add_button = tk.Button(root, text="Add Product", command=self.add_product)
        self.add_button.grid(row=4, column=0, columnspan=2)

        self.remove_button = tk.Button(root, text="Remove Product", command=self.remove_product)
        self.remove_button.grid(row=5, column=0, columnspan=2)

        self.view_button = tk.Button(root, text="View Inventory", command=self.view_inventory)
        self.view_button.grid(row=6, column=0, columnspan=2)

        self.inventory_text = tk.Text(root, height=10, width=40)
        self.inventory_text.grid(row=7, column=0, columnspan=2)

    def add_product(self):
        product_id = self.product_id_entry.get()
        name = self.name_entry.get()
        price = float(self.price_entry.get())
        quantity = int(self.quantity_entry.get())
        self.supermarket.add_product(product_id, name, price, quantity)
        messagebox.showinfo("Success", "Product added to inventory.")

    def remove_product(self):
        product_id = self.product_id_entry.get()
        quantity = int(self.quantity_entry.get())
        self.supermarket.remove_product(product_id, quantity)
        messagebox.showinfo("Success", "Product removed from inventory.")

    def view_inventory(self):
        inventory = self.supermarket.view_inventory()
        self.inventory_text.delete(1.0, tk.END)
        for product_id, details in inventory.items():
            self.inventory_text.insert(tk.END, f"Product ID: {product_id}, Name: {details['name']}, Price: ${details['price']}, Quantity: {details['quantity']}\n")

if __name__ == "__main__":
    root = tk.Tk()
    app = InventoryManagementApp(root)
    root.mainloop()
